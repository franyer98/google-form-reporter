"""
Google Form -> Excel Report Generator
--------------------------------------
Lee las respuestas de un formulario de Google (a través de la hoja de
cálculo vinculada) y genera un reporte en Excel con:
  - Hoja "Datos": respuestas crudas
  - Hoja "Resumen": conteos y estadísticas por pregunta
  - Gráficos de barras para las preguntas de tipo selección

Uso:
    python report_generator.py
"""

import base64
import io
import os
import tempfile
from datetime import datetime
from html import escape

import gspread
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from dotenv import load_dotenv
from google.oauth2.service_account import Credentials
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

load_dotenv()

SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]

SHEET_ID = os.getenv("SHEET_ID")
WORKSHEET_NAME = os.getenv("WORKSHEET_NAME", "Respuestas de formulario 1")
CREDENTIALS_PATH = os.getenv("GOOGLE_CREDENTIALS_PATH", "credentials.json")
OUTPUT_DIR = os.getenv("OUTPUT_DIR", "reportes")


def fetch_data() -> pd.DataFrame:
    """Conecta con Google Sheets y devuelve las respuestas como DataFrame."""
    if not SHEET_ID:
        raise ValueError("Falta la variable de entorno SHEET_ID.")
    if not os.path.exists(CREDENTIALS_PATH):
        raise FileNotFoundError(
            f"No se encontró el archivo de credenciales en '{CREDENTIALS_PATH}'."
        )

    creds = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
    client = gspread.authorize(creds)

    try:
        spreadsheet = client.open_by_key(SHEET_ID)
    except gspread.exceptions.SpreadsheetNotFound as exc:
        raise ValueError(
            f"No se encontró la hoja de cálculo con SHEET_ID='{SHEET_ID}'. "
            "Verificá el ID y que la cuenta de servicio tenga acceso de lector."
        ) from exc

    try:
        worksheet = spreadsheet.worksheet(WORKSHEET_NAME)
    except gspread.exceptions.WorksheetNotFound as exc:
        raise ValueError(
            f"No se encontró la pestaña '{WORKSHEET_NAME}' en la hoja de cálculo."
        ) from exc

    records = worksheet.get_all_records()
    if not records:
        raise ValueError("La hoja no tiene respuestas todavía.")

    return pd.DataFrame(records)


def build_summary(df: pd.DataFrame) -> dict:
    """Genera un resumen por columna: conteos para texto, estadísticas para números.

    Devuelve un dict {pregunta: {"type": "numeric"|"categorical", "data": Series}}
    """
    summary = {}
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            summary[col] = {"type": "numeric", "data": df[col].describe().round(2)}
        else:
            summary[col] = {"type": "categorical", "data": df[col].value_counts()}
    return summary


def write_excel_report(df: pd.DataFrame, summary: dict, output_path: str) -> None:
    """Escribe el reporte en Excel con datos, resumen y gráficos."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Datos", index=False)

        start_row = 0
        summary_sheet_name = "Resumen"

        for question, info in summary.items():
            result = info["data"]
            frame = result.reset_index()
            frame.columns = [question, "Cantidad" if info["type"] == "categorical" else "Valor"]
            frame.to_excel(
                writer,
                sheet_name=summary_sheet_name,
                startrow=start_row,
                index=False,
            )
            start_row += len(frame) + 3

        writer.sheets[summary_sheet_name]

    # Segunda pasada: añadir gráficos de barras a las preguntas categóricas
    from openpyxl import load_workbook

    wb = load_workbook(output_path)
    ws = wb["Resumen"]

    row_cursor = 1
    chart_col = "E"
    chart_row = 1

    for question, info in summary.items():
        n_rows = len(info["data"])
        if info["type"] == "categorical" and n_rows > 0:
            # Es un conteo categórico -> graficar
            data_ref = Reference(
                ws,
                min_col=2,
                min_row=row_cursor,
                max_row=row_cursor + n_rows,
            )
            cats_ref = Reference(
                ws,
                min_col=1,
                min_row=row_cursor + 1,
                max_row=row_cursor + n_rows,
            )
            chart = BarChart()
            chart.title = question[:40]
            chart.y_axis.title = "Cantidad"
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.width = 14
            chart.height = 8
            ws.add_chart(chart, f"{chart_col}{chart_row}")
            chart_row += 17

        row_cursor += n_rows + 3

    # Ajustar ancho de columnas en la hoja de Datos
    ws_datos = wb["Datos"]
    for i, col in enumerate(df.columns, start=1):
        max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
        ws_datos.column_dimensions[get_column_letter(i)].width = min(max_len, 50)

    wb.save(output_path)


def write_pdf_report(df: pd.DataFrame, summary: dict, output_path: str) -> None:
    """Genera un PDF con resumen, gráficos y la tabla completa de datos."""
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Reporte de Formulario", styles["Title"]))
    story.append(
        Paragraph(
            f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} — "
            f"Total de respuestas: {len(df)}",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 16))

    story.append(Paragraph("Resumen por pregunta", styles["Heading1"]))

    with tempfile.TemporaryDirectory() as tmp_dir:
        for question, info in summary.items():
            result = info["data"]
            if len(result) == 0:
                continue

            story.append(Spacer(1, 10))
            story.append(Paragraph(question, styles["Heading2"]))

            if info["type"] == "categorical":
                # Gráfico de barras
                fig, ax = plt.subplots(figsize=(6, 3))
                result.plot(kind="bar", ax=ax, color="#4C72B0")
                ax.set_ylabel("Cantidad")
                ax.set_xlabel("")
                plt.xticks(rotation=30, ha="right")
                plt.tight_layout()

                img_path = os.path.join(tmp_dir, f"{abs(hash(question))}.png")
                fig.savefig(img_path, dpi=150)
                plt.close(fig)

                story.append(Image(img_path, width=5.5 * inch, height=2.5 * inch))
            else:
                # Tabla de estadísticas numéricas
                stats_data = [["Estadística", "Valor"]] + [
                    [str(idx), str(val)] for idx, val in result.items()
                ]
                stats_table = Table(stats_data, colWidths=[150, 150])
                stats_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4C72B0")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                            ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ]
                    )
                )
                story.append(stats_table)

        story.append(PageBreak())

        # Tabla completa de datos crudos
        story.append(Paragraph("Datos completos", styles["Heading1"]))
        story.append(Spacer(1, 8))

        table_data = [list(df.columns)] + df.astype(str).values.tolist()
        col_width = max(0.9 * inch, (7.5 * inch) / max(len(df.columns), 1))
        data_table = Table(table_data, colWidths=col_width, repeatRows=1)
        data_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#333333")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
                ]
            )
        )
        story.append(data_table)

        doc = SimpleDocTemplate(output_path, pagesize=letter, topMargin=40, bottomMargin=40)
        doc.build(story)


def _chart_to_base64(result: pd.Series, question: str) -> str:
    """Renderiza un gráfico de barras y lo devuelve como PNG codificado en base64."""
    fig, ax = plt.subplots(figsize=(6, 3))
    result.plot(kind="bar", ax=ax, color="#4C72B0")
    ax.set_ylabel("Cantidad")
    ax.set_xlabel("")
    plt.xticks(rotation=30, ha="right")
    plt.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150)
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("ascii")


def write_html_report(df: pd.DataFrame, summary: dict, output_path: str) -> None:
    """Genera una página HTML autocontenida para ver el reporte en el navegador, sin descargar nada."""
    sections = []
    for question, info in summary.items():
        result = info["data"]
        if len(result) == 0:
            continue

        safe_question = escape(str(question))

        if info["type"] == "categorical":
            img_b64 = _chart_to_base64(result, question)
            body = f'<img src="data:image/png;base64,{img_b64}" alt="{safe_question}">'
        else:
            rows = "".join(
                f"<tr><td>{escape(str(idx))}</td><td>{escape(str(val))}</td></tr>"
                for idx, val in result.items()
            )
            body = (
                "<table class='stats'>"
                "<tr><th>Estadística</th><th>Valor</th></tr>"
                f"{rows}</table>"
            )

        sections.append(f"<section><h2>{safe_question}</h2>{body}</section>")

    data_header = "".join(f"<th>{escape(str(col))}</th>" for col in df.columns)
    data_rows = "".join(
        "<tr>" + "".join(f"<td>{escape(str(v))}</td>" for v in row) + "</tr>"
        for row in df.astype(str).values.tolist()
    )

    html = f"""<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Reporte de Formulario</title>
<style>
  :root {{ color-scheme: light dark; }}
  body {{
    font-family: -apple-system, "Segoe UI", Roboto, sans-serif;
    margin: 0 auto;
    padding: 2rem;
    max-width: 1000px;
  }}
  h1 {{ margin-bottom: 0.25rem; }}
  .meta {{ color: #767676; margin-bottom: 2rem; }}
  section {{ margin-bottom: 2.5rem; }}
  section h2 {{ font-size: 1.1rem; border-bottom: 1px solid #8884; padding-bottom: 0.4rem; }}
  img {{ max-width: 100%; height: auto; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 0.85rem; }}
  th, td {{ border: 1px solid #8884; padding: 0.4rem 0.6rem; text-align: left; }}
  th {{ background: #333; color: #fff; position: sticky; top: 0; }}
  table.stats th {{ background: #4C72B0; }}
  .data-wrap {{ overflow-x: auto; max-height: 600px; border: 1px solid #8884; }}
  tbody tr:nth-child(even) {{ background: rgba(128, 128, 128, 0.08); }}
</style>
</head>
<body>
  <h1>Reporte de Formulario</h1>
  <p class="meta">Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} — Total de respuestas: {len(df)}</p>

  <h2>Resumen por pregunta</h2>
  {''.join(sections)}

  <h2>Datos completos</h2>
  <div class="data-wrap">
    <table>
      <thead><tr>{data_header}</tr></thead>
      <tbody>{data_rows}</tbody>
    </table>
  </div>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs("site", exist_ok=True)

    df = fetch_data()
    summary = build_summary(df)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    excel_path = os.path.join(OUTPUT_DIR, f"reporte_{timestamp}.xlsx")
    pdf_path = os.path.join(OUTPUT_DIR, f"reporte_{timestamp}.pdf")
    html_path = os.path.join("site", "index.html")

    write_excel_report(df, summary, excel_path)
    write_pdf_report(df, summary, pdf_path)
    write_html_report(df, summary, html_path)

    print(f"Reporte Excel generado: {excel_path}")
    print(f"Reporte PDF generado: {pdf_path}")
    print(f"Reporte HTML generado: {html_path}")
    print(f"Total de respuestas: {len(df)}")


if __name__ == "__main__":
    main()
